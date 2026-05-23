from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def get_data(
            file_path,
            sheet_name,
            row,
            column
    ):

        workbook = load_workbook(
            file_path
        )

        sheet = workbook[
            sheet_name
        ]

        return sheet.cell(
            row=row,
            column=column
        ).value